import wx
import ProjectGUI
import configparser
import pathlib
import time
import hashlib
import threading
import math
from os import path

#############################
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

# Object for each order in Excel file
class CutOrder:

    def __init__(self, id, due_date, priority, vars_dict):
        self._id = id
        self._due_date = due_date
        self._priority = 1 if priority == 'High' else 0
        self._order_vars_dict = vars_dict
        self._oct = 0
        self._ost = 0
        self._tdt = 0

    def do_calcs(self, const_dict):
        # calculate OST, OCT, and TDT and update them
        # OST:
        MMT = const_dict['const_PST'] * self._order_vars_dict['spr_NM'] * 2
        MNT = const_dict['const_MST'] * self._order_vars_dict['spr_NM']
        MUT = const_dict['const_PST'] * self._order_vars_dict['spr_NM']
        MS = self._order_vars_dict['spr_TCY'] / const_dict['const_SS']
        MT = self._order_vars_dict['spr_TCY'] / (const_dict['const_ST'] * const_dict['const_STF'])
        MRT = (self._order_vars_dict['spr_TCL'] * 0.5) * (self._order_vars_dict['spr_TNR']) / const_dict['const_ST']
        MLT = const_dict['const_MLR'] * (self._order_vars_dict['spr_TNR'] - 1)
        MCT = const_dict['const_CRT'] * const_dict['const_CRF']
        MDT = self._order_vars_dict['spr_TCY'] / (self._order_vars_dict['spr_DY'] * const_dict['const_DT'])
        XSST = (const_dict['const_CST'] + MMT + MNT + MUT + const_dict['const_SSA']) / const_dict['const_OEF']
        XST = (MS + MT + MRT + MLT + MCT + MDT) / const_dict['const_OEF']
        self._ost = XSST + XST

        # OCT:
        MCMT2 = self._order_vars_dict['cut_TCL'] / const_dict['const_CV']
        NCM2 = self._order_vars_dict['cut_TCL'] / const_dict['const_CLT']
        MCMTS2 = NCM2 * const_dict['const_CLTS']
        MCT2 = self._order_vars_dict['cut_TCP1'] / const_dict['const_CS']
        MPMT2 = self._order_vars_dict['cut_TCP2'] * const_dict['const_PMT']
        self._oct = (MCMT2 + MCMTS2 + MCT2 + MPMT2 + const_dict['const_CST']) / const_dict['const_OEF']

        self._tdt = self._oct + self._ost

        #print('calculations finished!')

    def get_oct(self):
        return self._oct

    def get_ost(self):
        return self._ost

    def get_tdt(self):
        return self._tdt

    def __str__(self):
        return ("Sample object:\n"
                "  Name = {0}\n"
                "  Due Date = {1}\n"
                "  Priority = {2}\n"
                "  Total Time = {3}"
                .format(self._id, self._due_date, self._priority,
                    self._tdt))

#################################

processing = False

class SeniorProject(ProjectGUI.Dialog):
   def OnClose(self, event):
      if not processing:
         self.Destroy()
         exit()

   def OnExit(self, event):
      self.OnClose(event)

   def pause(self, timeout):
      if (timeout > 0):
         loop_count = int(timeout/100)
         remainder = timeout/100 - loop_count
         for count in range(loop_count):
            # yield to allow GUI application to remain responsive while timing
            wx.Yield()
            time.sleep(0.1)
            wx.Yield()
         time.sleep(remainder)

   def md5_file(self, file, blocksize=4096):
      with open(file, "rb") as f:
         file_hash = hashlib.md5()
         for block in iter(lambda: f.read(blocksize), b""):
            file_hash.update(block)
      return file_hash.hexdigest()

   def processFile(self, file):
      self.m_status_message.SetLabel("Processing " + str(file))
      print("Processing " + str(file))
      # processFile currently performs md5 on file
      # replace with desired processing
      self.pause(100)  #artifical pause (to be removed)
      md5 = self.md5_file(file)
      self.m_status_message.SetLabel("md5 of " + str(file) + " = " + str(md5))
      print("Finished. md5 of " + str(file) + " = " + str(md5))
      self.pause(100)  #artifical pause (to be removed)

   def enable_buttons(self, value):
      self.m_exit_button.Enable(value)
      self.m_help_button.Enable(value)
      self.m_save_button.Enable(value)
      #self.m_input_dir.Enable(value)
      self.m_output_dir.Enable(value)
      #self.m_input_extension.Enable(value)
      # change text on Start button to Stop when processing
      if value:
         self.m_start_button.SetLabel("Start")
      else:
         self.m_start_button.SetLabel("Stop")

   """def processDirectory(self, directory, type):
      global processing
      if path.exists(directory):
         self.enable_buttons(False)
         for file in pathlib.Path(directory).glob(type):
            self.processFile(file)
            # if 'Stop' button pressed, exit loop
            if not processing:
               break
      if processing:
         processing = False
         self.m_status_message.SetLabel("Waiting...")
      else:
         self.m_status_message.SetLabel("Stopped...")
      self.enable_buttons(True)"""
######################
# This is where the work needs to be done
   def OnStart(self, event):
      """global processing
      # if not already processing
      if not processing:
         processing = True
         #self.processDirectory(self.m_input_dir.GetPath(),self.m_input_extension.GetValue())
         x = threading.Thread(target=self.processDirectory, args=(self.m_input_dir.GetPath(), self.m_input_extension.GetValue(),), daemon=True)
         x.start()
      else:
         processing = False"""
      cut_file = self.m_input_file.GetPath()
      const_dict = {}
      cut_orders = []

      wb = openpyxl.load_workbook(filename=cut_file, read_only=True, data_only=True)

      print(wb.sheetnames)
      # Set worksheet to workbook's active sheet
      ws = wb['constants']
      for row in ws.iter_rows(min_row=2, values_only=True):
          parameters = []
          for cell in row:
              if cell is not None:
                  parameters.append(cell)
          if (len(parameters) >=  3) and (isinstance(parameters[2], float) or (isinstance(parameters[2], int))):
              const_dict["const_" + parameters[0]] = parameters[2]

      print(const_dict)

      ws = wb['orders']
      for row in ws.iter_rows(min_row=2, values_only=True):
          order_dict = {}
          parameters = []
          for cell in row:
              if cell is not None:
                  parameters.append(cell)
          if (len(parameters) >= 12) and (isinstance(parameters[4], float) or (isinstance(parameters[4], int))):
              order_dict["spr_DY"] = parameters[4]
              order_dict["spr_NM"] = parameters[5]
              order_dict["spr_TNR"] = parameters[6]
              order_dict["spr_TCL"] = parameters[7]
              order_dict["spr_TCY"] = parameters[8]
              order_dict["cut_TCP1"] = parameters[9]
              order_dict["cut_TCP2"] = parameters[10]
              order_dict["cut_TCL"] = parameters[11]
              print(parameters[0], order_dict)

              cut_orders.append(CutOrder(parameters[0], parameters[1], parameters[2], order_dict))

      print(cut_orders)

      for order in cut_orders:
          order.do_calcs(const_dict)
          # print(order._id)
          # print(order.get_oct())
          # print(order.get_ost())
          # print(order.get_tdt())

      #cut_orders.sort(key=lambda x: (x._due_date, -x._priority, -x._tdt), reverse=False)


      # Assume due date and priority number takes priority over filling up all space on tables

      order_list = cut_orders
      # Take input for table lengths.  for now assume 3 of length 100
      # table_lengths = [200, 200, 200]
      table_lengths = [float(int(self.m_textCtrl3.GetValue())), float(int(self.m_textCtrl4.GetValue())), float(int(self.m_textCtrl5.GetValue()))]

      # time_to_move = 15 * 60
      time_to_move = float(int(self.m_textCtrl51.GetValue())) * 60
      current_table_iter = 0  # Assume start at first table aka position 0 - assume for now, but could
      #                         take this as an input later if we really wanted to
      # sort order list by due date, priority number, time to cut
      # Basically just use sorting from earlier example code
      cut_orders.sort(key=lambda x: (x._due_date, -x._priority, -x._tdt), reverse=False)
      for x in cut_orders:
          print(x._id, x._due_date, x._priority, x._tdt)

      # Divide into sublists - each sublist should have orders of the same due date and priority number, but
      # possibly different cut times
      # Start iterator at first position aka 0
      i = 0
      # start with empty queue:
      queue = []
      # and new order list to add sorted sublists to
      new_order_list = []
      order_left_to_process = len(order_list)
      while i < len(order_list):
          print("Current order index:", i)
          due_date = order_list[i]._due_date
          priority_num = order_list[i]._priority
          sublist = []
          sublist.append([order_list[i], 0])
          i += 1
          i_other_index = i
          for j in range(i_other_index, len(order_list)):
              # check each
              if order_list[j]._due_date == due_date and order_list[j]._priority == priority_num:
                  # Append order, along with spot to set as calculated wait / "in-between" time when sorting further down
                  # After further work, sublist doesn't actually have to be a list of lists
                  # So, the spot for in between time can be removed - but this will require
                  # Reformatting any reference to objects in the sublist
                  sublist.append([order_list[j], 0])
                  # del order_list[j]
              else:
                  print("Next order indexing should start at:", i)
                  break
                  # break upon finding orders with different priority or due date
              i += 1  # increment i so we will start from the next batch upon finding
              #         orders with different priority or due date

          # begin placing orders in queue
          # continue going through sublist to find next order until we have emptied the list
          while sublist:
              print("Parsing through current sublist")
              # Possible improvement: select first order to be spread as one that will take least spreading time
              # (and most cutting time if multiple have the least spread time)
              # Only do so if queue is empty
              if not queue:
                  first_order = None
                  while first_order is None:
                      print("Adding first order to empty queue")

                      for x in sublist:
                          if (((first_order is None) or (x[0]._ost < first_order[0]._ost and
                                                        x[0]._oct > first_order[0]._oct)) and
                                                        table_lengths[current_table_iter] >= x[0]._order_vars_dict['spr_TCL']):
                              first_order = x
                      print("First order found")
                      print(first_order, table_lengths[current_table_iter], first_order[0]._order_vars_dict['spr_TCL'])
                  # add found first order to queue
                  queue.append([first_order[0], first_order[0]._oct, current_table_iter])
                  # remove added order from sublist, add to new order list
                  sublist.remove(first_order)
                  new_order_list.append([first_order[0], current_table_iter])
                  order_left_to_process -= 1
                  # subtract length of first order from table being spread on
                  table_lengths[current_table_iter] -= first_order[0]._order_vars_dict['spr_TCL']
                  print("First order found and added to queue, proceeding with scheduling")
              # As of now, above assume any empty table will be large enough to accommodate any one order

              next_order = None
              next_order_table_index = None
              min_cutter_downtime = float('inf')
              max_cut_time = 0
              min_spread_wait_time = float('inf')

              for x in sublist:
                  # calculate current table wait times
                  table_cut_wait_times = [0, 0, 0]  # only used for finding cutter down times
                  table_cutter_move_wait_times = [0, 0, 0]  # only used for finding cutter down times
                  last_table_ind = None
                  # Calculate cut wait times and cutter move wait times for each table
                  for order in queue:
                      if last_table_ind is None:
                          table_cut_wait_times[order[2]] += order[1]
                          last_table_ind = order[2]
                      elif last_table_ind == order[2]:
                          table_cut_wait_times[order[2]] += order[1]
                      else:
                          table_cut_wait_times[order[2]] += order[1]
                          table_cutter_move_wait_times[order[2]] += time_to_move
                          last_table_ind = order[2]
                  # Calculate spread wait times for each table
                  table_spread_wait_times = [0, 0, 0]
                  table_lengths_copy = table_lengths
                  # for i in range(0, len(table_lengths)):
                  #    table_lengths_copy[i] = table_lengths[i]
                  for ind in range(0, len(table_lengths_copy)):
                      if table_lengths_copy[ind] < x[0]._order_vars_dict['spr_TCL']:
                          needed_length = x[0]._order_vars_dict['spr_TCL']
                          length_avail = table_lengths_copy[ind]
                          for y in queue:
                              if length_avail < needed_length:
                                  if y[2] == ind:
                                      table_spread_wait_times[ind] += y[1]
                                      length_avail += y[0]._order_vars_dict['spr_TCL']
                              else:
                                  break

                  # At this point we should now have:
                  # - cut wait times for each table, if any
                  # - cutter moving time for each table
                  # - wait times for spreading for each table, for this specific order
                  # We want to minimize cutter down time and spreading wait time
                  # Minimizing cutter down time will take priority
                  cur_cutter_downtimes = [0, 0, 0]
                  for i_2 in range(0, len(cur_cutter_downtimes)):
                      cur_cutter_downtimes[i_2] = table_spread_wait_times[i_2] \
                                                - table_cut_wait_times[i_2] \
                                                + table_cutter_move_wait_times[i_2]
                      if cur_cutter_downtimes[i_2] < 0:
                          cur_cutter_downtimes[i_2] = 0
                  # In order of precedence:
                  # 1. Minimize cutter down time
                  # 2. Maximize order cut time (per original priority instructions)
                  # 3. Minimize spread wait time
                  # We now find which table, for this specific order, would best meet these requirements
                  order_min_cutter_downtime = float('inf')
                  order_cut_time = x[0]._oct
                  order_min_spread_wait_time = float('inf')
                  table_index = None
                  # print("Line 326: calculating individual order time values")
                  for i_3 in range(0, len(table_lengths)):
                      if cur_cutter_downtimes[i_3] > order_min_cutter_downtime:
                          pass
                      elif cur_cutter_downtimes[i_3] == order_min_cutter_downtime:
                          if table_spread_wait_times[i_3] >= order_min_spread_wait_time:
                              pass
                          else:
                              order_min_spread_wait_time = table_spread_wait_times[i_3]
                              table_index = i_3
                              order_min_cutter_downtime = cur_cutter_downtimes[i_3]
                      elif cur_cutter_downtimes[i_3] < order_min_cutter_downtime:
                          order_min_spread_wait_time = table_spread_wait_times[i_3]
                          table_index = i_3
                          order_min_cutter_downtime = cur_cutter_downtimes[i_3]
                  # after this, we compare to the current values for the next order to be
                  # Again, if this order better meets the requirements for the current next order, we replace the current
                  # With this new one
                  # print("Line 344: comparing this order to current next order")
                  if order_min_cutter_downtime > min_cutter_downtime:
                      pass
                  elif order_min_cutter_downtime == min_cutter_downtime:
                      if order_cut_time < max_cut_time:
                          pass
                      elif order_cut_time == max_cut_time:
                          if order_min_spread_wait_time >= min_spread_wait_time:
                              pass
                          elif order_min_spread_wait_time < min_spread_wait_time:
                              next_order = x[0]
                              next_order_table_index = table_index
                              min_cutter_downtime = order_min_cutter_downtime
                              max_cut_time = order_cut_time
                              min_spread_wait_time = order_min_spread_wait_time
                      elif order_cut_time > max_cut_time:
                          next_order = x[0]
                          next_order_table_index = table_index
                          min_cutter_downtime = order_min_cutter_downtime
                          max_cut_time = order_cut_time
                          min_spread_wait_time = order_min_spread_wait_time
                  elif order_min_cutter_downtime < min_cutter_downtime:
                      next_order = x[0]
                      next_order_table_index = table_index
                      min_cutter_downtime = order_min_cutter_downtime
                      max_cut_time = order_cut_time
                      min_spread_wait_time = order_min_spread_wait_time

              # The above will occur for every order in the sublist
              # By the time it is done, we should have the order that has:
              # - The least cutter down time
              # - The longest time required to actually cut the order
              # - The minimal spreading wait time
              # We can now add this order to the end of the new order list
              # After doing so we will have to see if other orders have been completed and,
              # if so, remove them from the queue so it will reflect the state of the tables when
              # We look for the next order that should be spread

              # Add to new order list with table index
              new_order_list.append([next_order, next_order_table_index])
              order_left_to_process -= 1
              # Subtract length from table order is being added to
              table_lengths[next_order_table_index] -= next_order._order_vars_dict['spr_TCL']
              # Add to queue
              queue.append([next_order, next_order._oct, next_order_table_index])
              # remove from sublist
              #if next_order in sublist:
              to_remove = []
              print("Sublist pre-removal:")
              print(sublist)
              for l in sublist:
                if l[0] == next_order:
                    to_remove = l
              sublist.remove(to_remove)
              # print("Line 393: Item removed from sublist")
              print("Sublist post-removal:")
              print(sublist)
              # subtract spreading time of this order from oldest order in queue if there is one
              if len(queue) > 1:
                  queue[0] = [queue[0][0], queue[0][1] - (next_order._ost + min_spread_wait_time), queue[0][2]]
                  # if resulting time is negative, remove from queue and subtract (add negative) leftover time
                  # from next oldest order if next oldest order is not the new order
                  if queue[0][1] <= 0:
                      # new order check
                      if queue[1][0] == next_order:
                          table_lengths[queue[0][2]] += queue[0][0]._order_vars_dict['spr_TCL']
                          print(queue[0])
                          queue.pop(0)
                      else:
                          while queue[0][1] <= 0:
                              print("subtrtacting leftover time")

                              if queue[0][2] == queue[1][2]:
                                  queue[1][1] += queue[0][1]
                                  # also re-add length to table that is now being cleared
                                  table_lengths[queue[0][2]] += queue[0][0]._order_vars_dict['spr_TCL']
                                  queue.pop(0)
                              else:
                                  queue[1][1] += queue[0][1] + time_to_move
                                  # also re-add length to table that is now being cleared
                                  table_lengths[queue[0][2]] += queue[0][0]._order_vars_dict['spr_TCL']
                                  queue.pop(0)

      # Export new_order_list
      # for now, just print
      print(new_order_list)

      """namelist = []
      for order in cut_orders:
          namelist.append([order._id, math.floor(order._tdt/60)])"""

      namelist = []
      for order in new_order_list:
          namelist.append([order[0]._id, math.floor(order[0]._tdt / 60), order[1]])
      print("resulting schedule:")
      print(namelist)


      wb.close()

      # Create new excel file for ordered list of cut orders
      # Create new workbook
      wb = Workbook()
      # Change sheet to workbook's active sheet
      ws = wb.create_sheet("Ordered Cut List", 0)
        # Set Column headers
      ws['A1'] = 'Order ID'
      ws['B1'] = 'Required Date'
      ws['C1'] = 'Priority'
      ws['D1'] = 'Total Time'
      ws['E1'] = 'Table Index'

      # Go through sorted list of items, setting cell values from CutOrder object parameters
      row_iter = 2
      while row_iter < (len(new_order_list) + 1):
          for order in new_order_list:
              ws['A' + str(row_iter)] = order[0]._id
              ws['B' + str(row_iter)] = order[0]._due_date
              ws['C' + str(row_iter)] = order[0]._priority
              ws['D' + str(row_iter)] = math.floor(order[0]._tdt/60)
              ws['E' + str(row_iter)] = order[1]
              row_iter += 1

        # Save to new excel file
      #wb.save(self.m_output_dir.GetPath() + '\Ordered Cuts.xlsx')
      wb.save(self.m_output_dir.GetPath() + '\\' + str(self.m_textCtrl1.GetValue()) + '.xlsx')

      # Create new excel file for ordered list of cut orders
      # Create new workbook
      # wb = Workbook()
      # Change sheet to workbook's active sheet
      """ws = wb.create_sheet("Ordered Cut List", 0)
        # Set Column headers
      ws['A1'] = 'Order ID'
      ws['B1'] = 'Required Date'
      ws['C1'] = 'Priority'
      ws['D1'] = 'Total Time'
      
      # Go through sorted list of items, setting cell values from CutOrder object parameters
      row_iter = 2
      while row_iter < (len(cut_orders) + 1):
          for order in cut_orders:
              ws['A' + str(row_iter)] = order._id
              ws['B' + str(row_iter)] = order._due_date
              ws['C' + str(row_iter)] = order._priority
              ws['D' + str(row_iter)] = math.floor(order._tdt/60)
              row_iter += 1
      
        # Save to new excel file
      wb.save(self.m_output_dir.GetPath() + '\Ordered Cuts.xlsx')"""

################

   def OnSave(self, event):
      config = configparser.ConfigParser()
      config.add_section('configuration')
      config.set('configuration', 'output_dir', self.m_output_dir.GetPath())
      # config.set('configuration', 'input_dir', self.m_input_dir.GetPath())
      # config.set('configuration', 'input_extension', self.m_input_extension.GetValue())
      config.set('configuration', 'input_file', self.m_input_file.GetPath())
      config.set('configuration', 'output_file', str(self.m_textCtrl1.GetValue()))
      config.set('configuration', 'Table 0 Length', str(self.m_textCtrl3.GetValue()))
      config.set('configuration', 'Table 1 Length', str(self.m_textCtrl4.GetValue()))
      config.set('configuration', 'Table 2 Length', str(self.m_textCtrl5.GetValue()))
      config.set('configuration', 'Cutter Move Time', str(self.m_textCtrl51.GetValue()))

      with open('settings.ini', 'w') as configfile:
         config.write(configfile)

   def OnHelp(self, event):
      wx.MessageBox("Help for application", "Help", wx.ICON_INFORMATION)

   def OnLoad(self):
      if path.exists('settings.ini'):
         config = configparser.ConfigParser()
         config.read('settings.ini')
         self.m_output_dir.SetPath(config.get('configuration', 'output_dir', fallback=''))
         #self.m_input_dir.SetPath(config.get('configuration', 'input_dir', fallback=''))
         #self.m_input_extension.SetValue(config.get('configuration', 'input_extension', fallback='*.txt'))
         self.m_input_file.SetPath(config.get('configuration', 'input_file', fallback=''))
         self.m_textCtrl1.write(config.get('configuration', 'output_file', fallback=''))
         self.m_textCtrl3.write(config.get('configuration', 'Table 0 Length', fallback=''))
         self.m_textCtrl4.write(config.get('configuration', 'Table 1 Length', fallback=''))
         self.m_textCtrl5.write(config.get('configuration', 'Table 2 Length', fallback=''))
         self.m_textCtrl51.write(config.get('configuration', 'Cutter Move Time', fallback=''))


def main():
      # create a wx app, False stands for no redirection of stdin/stdout
      app = wx.App(False)
      # create an object of wxBacCmd
      frame = SeniorProject(None)
      # load the initial settings
      SeniorProject.OnLoad(frame)
      # show the frame
      frame.Show(True)
      # start the application
      app.MainLoop()

if __name__ == "__main__":
   main()
